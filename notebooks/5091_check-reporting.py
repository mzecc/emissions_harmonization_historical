# ---
# jupyter:
#   jupytext:
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.16.6
#   kernelspec:
#     display_name: Python 3 (ipykernel)
#     language: python
#     name: python3
# ---

# %% [markdown]
# # Model reporting
#
# Here we check the reporting of a given model.

# %% [markdown]
# ## Imports

# %%
from functools import partial

import gcages.cmip7_scenariomip.pre_processing.reaggregation.basic
import openpyxl.workbook.workbook
import pandas as pd
import pandas.io.excel
import pandas_indexing as pix
import tqdm.auto
from gcages.completeness import get_missing_levels
from openpyxl.styles.fonts import Font
from pandas_openscm.indexing import multi_index_match

from emissions_harmonization_historical.constants_5000 import (
    DATA_ROOT,
    DOWNLOAD_SCENARIOS_ID,
    RAW_SCENARIO_DB,
)
from emissions_harmonization_historical.excel_writing import set_cell

# %% [markdown]
# ## Set up

# %% editable=true slideshow={"slide_type": ""} tags=["parameters"]
model: str = "MESSAGE"

# %%
output_dir_model = DATA_ROOT / "raw" / "scenarios" / DOWNLOAD_SCENARIOS_ID / model
output_dir_model.mkdir(exist_ok=True, parents=True)
output_dir_model

# %%
out_file = output_dir_model / f"reporting-checking_{model}.xlsx"
out_file

# %%
pd.set_option("display.max_colwidth", None)

# %% [markdown]
# ## Load data

# %% [markdown]
# ### Scenarios

# %%
model_raw = RAW_SCENARIO_DB.load(pix.ismatch(model=f"*{model}*"), progress=True)
if model_raw.empty:
    raise AssertionError

# model_raw

# %% [markdown]
# ### Properties

# %%
properties = pd.concat([pd.read_csv(f) for f in output_dir_model.glob("*properties.csv")]).drop(
    "metadata", axis="columns"
)
# properties

# %% [markdown]
# ## Check completeness

# %% [markdown]
# Extract the model data, keeping:
#
# - only reported timesteps
# - only data from 2015 onwards (we don't care about data before this)
# - only data up to 2100 (while reporting is still weird for data post 2100)

# %%
model_df = model_raw.loc[:, 2015:2100].dropna(how="all", axis="columns")
if model_df.empty:
    raise AssertionError

model_df.columns.name = "year"
# model_df

# %% [markdown]
# ### Figure out the model-specific regions

# %%
model_regions = [r for r in model_df.pix.unique("region") if r.startswith(model.split(" ")[0])]
if not model_regions:
    raise AssertionError

# model_regions

# %% [markdown]
# ### Define the required timeseries index
#
# If the model reports domestic aviation at the regional level, this will be fine.
# If not, we'll need to add some different code.

# %%
required_timeseries_index = gcages.cmip7_scenariomip.pre_processing.reaggregation.basic.get_required_timeseries_index(
    model_regions
)
# required_timeseries_index.to_frame(index=False)

# %% [markdown]
# ### Get the missing timeseries

# %%
missing_l = []
for (model_l, scenario), msdf in tqdm.auto.tqdm(model_df.groupby(["model", "scenario"])):
    missing_df = get_missing_levels(msdf.index, required_timeseries_index, unit_col="unit").to_frame(index=False)
    missing_df["model"] = model_l
    missing_df["scenario"] = scenario

    missing_l.append(missing_df)

missing = pix.concat(missing_l)[["model", "scenario", "region", "variable"]]
missing


# %% [markdown]
# ### Report the missing timeseries

# %% [markdown]
# #### Helper functions


# %%
def get_common_missing_values(missing: pd.DataFrame, cols_of_interest: list[str]) -> pd.DataFrame:
    """
    Get missing values that are common for the columns of interest
    """
    # Probably a slow way to do this, but speed not an issue now
    missing_across_all_facets = None
    group_facets = missing.columns.difference(cols_of_interest).tolist()
    for _, msdf in missing.groupby(group_facets):
        if missing_across_all_facets is None:
            missing_across_all_facets = msdf[cols_of_interest]
        else:
            missing_across_all_facets = pd.merge(
                msdf[cols_of_interest], missing_across_all_facets, how="inner", on=cols_of_interest
            )

    if missing_across_all_facets is None:
        # return empty df
        return missing.iloc[:0, :]

    return missing_across_all_facets


# %%
def get_missing_across_facets(missing: pd.DataFrame, group_facets: list[str]) -> pd.DataFrame:
    """
    Get missing values that are common across the given facets
    """
    non_groupers = missing.columns.difference(group_facets).tolist()
    return get_common_missing_values(missing, non_groupers)


# %%
get_missing_across_scenarios = partial(get_missing_across_facets, group_facets=["model", "scenario"])


# %%
def get_variants_not_common(missing: pd.DataFrame, missing_common: pd.DataFrame):
    """
    Get the variations in missing that have not been identified as being common across some facets
    """
    tmp = missing.set_index(missing_common.columns.tolist())
    if not isinstance(tmp.index, pd.MultiIndex):
        tmp.index = pd.MultiIndex.from_product([tmp.index.values], names=[tmp.index.name])

    missing_common_idx = pd.MultiIndex.from_frame(missing_common)
    variants_not_common = tmp.loc[~multi_index_match(tmp.index, missing_common_idx)]

    return variants_not_common


# %%
def write_missing_sheet(  # noqa: PLR0913
    sheet_name: str,
    *,
    df: pd.DataFrame,
    first_row_comment: str,
    first_row_comment_if_empty: str,
    workbook: openpyxl.workbook.workbook.Workbook,
    writer: pandas.io.excel.ExcelWriter,
) -> None:
    """
    Write a sheet about missing data
    """
    if df.empty:
        set_cell(
            first_row_comment_if_empty,
            row=0,
            col=0,
            ws=workbook.create_sheet(title=sheet_name),
            font=Font(bold=True),
        )

    else:
        set_cell(
            first_row_comment,
            row=0,
            col=0,
            ws=workbook.create_sheet(title=sheet_name),
            font=Font(bold=True),
        )
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=2,
        )


# %% [markdown]
# #### Do the reporting

# %%
world_locator = missing["region"] == "World"
missing_world = missing[world_locator]
missing_model_region = missing[~world_locator]

# %%
with pd.ExcelWriter(out_file, engine="openpyxl", mode="w") as writer:
    workbook = writer.book
    sheet_name = "scenario_database_info"
    set_cell(
        "Scenario properties as downloaded from the database",
        row=0,
        col=0,
        ws=workbook.create_sheet(title=sheet_name),
        font=Font(bold=True),
    )
    properties.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=2,
        index=False,
    )

# %%
with pd.ExcelWriter(out_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    workbook = writer.book
    wms = partial(write_missing_sheet, workbook=workbook, writer=writer)
    wms(
        sheet_name="world",
        df=missing_world.reset_index(drop=True),
        first_row_comment="The following timeseries are missing at the World level",
        first_row_comment_if_empty="As far as we can tell, there are no timeseries missing at the World level",
    )
    if missing_world.empty:
        print("Nothing missing at the world level")

    missing_world_all_scenarios = get_missing_across_scenarios(missing_world)
    wms(
        sheet_name="world_all_scenarios",
        df=missing_world_all_scenarios.reset_index(drop=True),
        first_row_comment="The following timeseries are missing at the World level for all scenarios",
        first_row_comment_if_empty=(
            "As far as we can tell, there are no timeseries missing at the World level for all scenarios"
        ),
    )
    if not missing_world_all_scenarios.empty:
        print("The following timeseries are missing at the World level for all scenarios")
        display(missing_world_all_scenarios)  # noqa: F821

    missing_world_specific_scenarios = get_variants_not_common(missing_world, missing_world_all_scenarios)
    wms(
        sheet_name="world_specific_scenarios",
        df=missing_world_specific_scenarios,
        first_row_comment="The following timeseries are missing at the World level only for specific scenarios",
        first_row_comment_if_empty=(
            "As far as we can tell, there are no timeseries missing at the World level only for specific scenarios"
        ),
    )
    if not missing_world_specific_scenarios.empty:
        print("The following timeseries are missing at the World level only for specific scenarios")
        display(missing_world_specific_scenarios)  # noqa: F821

# %%
with pd.ExcelWriter(out_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    workbook = writer.book
    wms = partial(write_missing_sheet, workbook=workbook, writer=writer)

    wms(
        sheet_name="model_region",
        df=missing_model_region.reset_index(drop=True),
        first_row_comment="The following timeseries are missing at the model region level",
        first_row_comment_if_empty="As far as we can tell, there are no timeseries missing at the model region level",
    )
    if missing_model_region.empty:
        print("Nothing missing at the model-region level")

    else:
        missing_model_region_all_scenarios_regions = get_missing_across_facets(
            missing_model_region, group_facets=["model", "scenario", "region"]
        )
        wms(
            sheet_name="mr_all_scenarios_regions",
            df=missing_model_region_all_scenarios_regions.reset_index(drop=True),
            first_row_comment=(
                "The following timeseries are missing at the model region level for all scenarios and regions"
            ),
            first_row_comment_if_empty=(
                "As far as we can tell, "
                "there are no timeseries missing at the model region level for all scenarios and regions"
            ),
        )
        if not missing_model_region_all_scenarios_regions.empty:
            print("The following timeseries are missing at the model region level for all scenarios and regions")
            display(missing_model_region_all_scenarios_regions)  # noqa: F821

        missing_model_region_all_scenarios = get_variants_not_common(
            missing_model_region[["region", "variable"]].drop_duplicates(),
            missing_model_region_all_scenarios_regions,
        )
        wms(
            sheet_name="mr_all_scenarios",
            df=missing_model_region_all_scenarios.reset_index(),
            first_row_comment="The following timeseries are missing at the model region level for all scenarios",
            first_row_comment_if_empty=(
                "As far as we can tell, there are no timeseries missing at the model region level for all scenarios"
            ),
        )
        if not missing_model_region_all_scenarios.empty:
            print("The following timeseries are missing at the model region level for all scenarios")
            display(missing_model_region_all_scenarios)  # noqa: F821

        missing_model_region_specific_scenarios_regions = get_variants_not_common(
            missing_model_region,
            missing_model_region_all_scenarios,
        )
        missing_model_region_specific_scenarios_regions = get_variants_not_common(
            missing_model_region_specific_scenarios_regions,
            missing_model_region_all_scenarios_regions,
        )
        wms(
            sheet_name="mr_specific_region_scenarios",
            df=missing_model_region_specific_scenarios_regions.reset_index(drop=True),
            first_row_comment=(
                "The following timeseries are missing at the model region level for specific regions and scenarios"
            ),
            first_row_comment_if_empty=(
                "As far as we can tell, "
                "there are no timeseries missing at the model region level for specific regions and scenarios"
            ),
        )
        if not missing_model_region_specific_scenarios_regions.empty:
            print("The following timeseries are missing for specifc regions and scenarios")
            display(missing_model_region_specific_scenarios_regions)  # noqa: F821
